import os
import sys
import shutil
from pathlib import Path
from ultralytics import YOLO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import argparse

def delete_cache(path: Path):
    if path.exists():
        if path.is_file():
            path.unlink()
        elif path.is_dir():
            shutil.rmtree(path)

def process_image(model, image_path: Path, shared_output_dir: Path, flat_crop_dir: Path, detections_accumulator=None):
    results = model.predict(
        source=str(image_path),
        save=True,
        project=str(shared_output_dir.parent),
        name=shared_output_dir.name,
        exist_ok=True,
        save_crop=True
    )

    for r in results:
        for i, box in enumerate(r.boxes):
            cls_id = int(box.cls)
            cls_name = model.names[cls_id]
            conf = box.conf.item()
            xyxy = [round(x, 2) for x in box.xyxy.tolist()[0]]

            crop_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{cls_name}_{i+1}.jpg"
            crop_dir = shared_output_dir / "crops" / cls_name
            crop_file_candidates = list(crop_dir.glob("*.jpg"))
            crop_file = crop_file_candidates[i] if i < len(crop_file_candidates) else None

            flat_crop_path = flat_crop_dir / crop_name
            if crop_file and crop_file.is_file():
                shutil.copy(crop_file, flat_crop_path)

            if detections_accumulator is not None:
                detections_accumulator.append({
                    "file": image_path.name,
                    "class": cls_name,
                    "confidence": conf,
                    "box": xyxy,
                    "flat_crop_path": str(flat_crop_path)
                })

def export_to_excel(detections, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Detections"
    headers = ["Filename", "Class", "Confidence (%)", "Box x1", "Box y1", "Box x2", "Box y2", "Flat Crop Path"]
    ws.append(headers)

    for det in detections:
        ws.append([
            det["file"],
            det["class"],
            round(det["confidence"] * 100, 2),
            *det["box"],
            det.get("flat_crop_path", "")
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    wb.save(output_path)

def main():
    parser = argparse.ArgumentParser(description="YOLOv8 Batch Detector")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('--image', type=str, help="Path to a single image")
    group.add_argument('--folder', type=str, help="Path to folder of images")
    group.add_argument('--video', type=str, help="Path to video file")
    parser.add_argument('--model', type=str, default="yolov8x.pt", help="YOLOv8 model to use")
    args = parser.parse_args()

    model = YOLO(args.model)
    output_base = Path("yolo_output")
    shared_output_dir = output_base / "processed_images"
    flat_crop_dir = output_base / "flat_crops"
    shared_output_dir.mkdir(parents=True, exist_ok=True)
    flat_crop_dir.mkdir(parents=True, exist_ok=True)
    detections_log = []

    if args.image:
        image_path = Path(args.image)
        if not image_path.is_file():
            print(f"Error: Image {image_path} not found.")
            sys.exit(1)
        process_image(model, image_path, shared_output_dir, flat_crop_dir, detections_accumulator=detections_log)

    elif args.folder:
        folder_path = Path(args.folder)
        if not folder_path.is_dir():
            print(f"Error: Folder {folder_path} not found.")
            sys.exit(1)
        for img_file in folder_path.glob("*.[jp][pn]g"):
            process_image(model, img_file, shared_output_dir, flat_crop_dir, detections_accumulator=detections_log)

    elif args.video:
        video_path = Path(args.video)
        if not video_path.is_file():
            print(f"Error: Video {video_path} not found.")
            sys.exit(1)
        model.predict(
            source=str(video_path),
            save=True,
            project=str(shared_output_dir.parent),
            name=shared_output_dir.name,
            exist_ok=True
        )

    if detections_log:
        export_to_excel(detections_log, output_base / "detections_with_flat_crops.xlsx")

if __name__ == "__main__":
    main()
